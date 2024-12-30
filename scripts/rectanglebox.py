import cv2
import numpy as np
import random
from openpyxl import Workbook, load_workbook
import matplotlib.pyplot as plt
import os

# Load the output image with blue areas
input_image_path = os.path.join(os.path.dirname(__file__), '../book/static/book/images/cbf2024.jpg')  # Update with your output image path
image = cv2.imread(input_image_path)

if image is None:
    print("Error: Image not found. Please check the path.")
    exit()

# Load the Excel file to get the box data
excel_path = r'D:\Book-fair\bookstall\book\box_information_with_ocr.xlsx'  # Update with your Excel path
wb = load_workbook(excel_path)
ws = wb.active

# Extract box data from the Excel file
boxes = []
for row in ws.iter_rows(min_row=2, values_only=True):  # Skip the header row
    no, x, y, width, height, area, name = row[:7]
    if name is not None and x is not None and y is not None and width is not None and height is not None:
        name = str(name).strip()  # Convert 'name' to a string and strip extra spaces
        boxes.append((name, x, y, width, height))

# Determine bounds for colorful background
left_bound = min(x for _, x, _, _, _ in boxes)  # Left-most box boundary
right_bound = max(x + width for _, x, _, width, _ in boxes)  # Right-most box boundary
top_bound = min(y for _, _, y, _, _ in boxes)  # Top-most box boundary
bottom_bound = max(y + height for _, _, y, _, height in boxes)  # Bottom-most box boundary

# Create a mask where the boxes are placed (to avoid these areas)
mask = np.zeros_like(image, dtype=np.uint8)

for _, x, y, w, h in boxes:
    cv2.rectangle(mask, (int(x), int(y)), (int(x + w), int(y + h)), (255, 255, 255), -1)

# Function to generate a random color for the new squares/rectangles
def generate_random_color():
    return [random.randint(0, 255) for _ in range(3)]  # Generate random color (BGR format)

# Prepare a list to store new rectangles' coordinates and dimensions
new_rectangles = []

# Now, iterate over the image and fill non-box areas with random rectangles and squares
height, width, _ = image.shape
square_size = 35  # Set the size of the squares (or adjust as needed)

for i in range(0, height, square_size):
    for j in range(0, width, square_size):
        # Skip areas outside the bounds
        if j < left_bound or j >= right_bound or i < top_bound or i >= bottom_bound:
            continue

        # Define the square's top-left and bottom-right corners
        square_top_left = (j, i)
        square_bottom_right = (j + square_size, i + square_size)

        # Make sure the square doesn't go out of bounds
        if square_bottom_right[0] > width:
            square_bottom_right = (width, square_bottom_right[1])
        if square_bottom_right[1] > height:
            square_bottom_right = (square_bottom_right[0], height)

        # Check if this square overlaps with any box region
        if np.all(mask[i:square_bottom_right[1], j:square_bottom_right[0]] == 0):
            # If no overlap, draw a random colored square/rectangle
            color = generate_random_color()
            cv2.rectangle(image, square_top_left, square_bottom_right, color, -1)

            # Save the rectangle's index, coordinates, and dimensions
            x, y = square_top_left
            w = square_bottom_right[0] - x
            h = square_bottom_right[1] - y
            new_rectangles.append((len(new_rectangles) + 1, x, y, w, h))

             # Add a number inside the rectangle
            text = str(len(new_rectangles))
            font_scale = 0.4  # Smaller font size
            font_thickness = 1
            text_size = cv2.getTextSize(text, cv2.FONT_HERSHEY_SIMPLEX, font_scale, font_thickness)[0]
            text_x = x + (w - text_size[0]) // 2
            text_y = y + (h + text_size[1]) // 2
            cv2.putText(image, text, (text_x, text_y), cv2.FONT_HERSHEY_SIMPLEX, font_scale, (255, 255, 255), font_thickness)


# Save the resulting image with new rectangles/squares
output_image_path = os.path.join(os.path.dirname(__file__), '../book/static/book/images/new_5_CBF2024-layout.jpg')  # Update with your desired output path
cv2.imwrite(output_image_path, image)

# Save the new rectangles' data to a new Excel file
output_excel_path = r'D:\Book-fair\bookstall\book\box_5_information.xlsx'  # Update with your desired Excel output path
wb_new = Workbook()
ws_new = wb_new.active
ws_new.title = "New Rectangles"

# Write the header row
ws_new.append(["Index", "X", "Y", "Width", "Height"])

# Write the rectangles' data
for rect in new_rectangles:
    ws_new.append(rect)

# Save the new Excel file
wb_new.save(output_excel_path)

# Display the image with new rectangles/squares
plt.figure(figsize=(10, 8))
plt.imshow(cv2.cvtColor(image, cv2.COLOR_BGR2RGB))
plt.title("Stall Layout with Colorful Background")
plt.axis('off')
plt.show()

print(f"Image with new squares/rectangles saved to: {output_image_path}")
print(f"New rectangles' data saved to: {output_excel_path}")




