import cv2
import numpy as np
from openpyxl import load_workbook
import os

# Load the image
image_path = os.path.join(os.path.dirname(__file__), '../book/static/book/images/cbf-layout1.jpg')  # Update with your output image path
image = cv2.imread(image_path)

if image is None:
    print("Error: Image not found.")
    exit()

# Load the Excel file with box information
excel_path = r"D:\Project-5\bookstall\book\book_stall.xlsx"
wb = load_workbook(excel_path)
ws = wb.active

# Extract box data from the Excel file
boxes = []
for row in ws.iter_rows(min_row=2, values_only=True):  # Skip the header row
    # Extract the first 5 columns only
    box_data = row[:5]
    if len(box_data) == 5:  # Ensure the row has the expected number of columns
        boxes.append(box_data)

# Sort boxes by the 'y' coordinate (top-to-bottom)
# Assuming box data is in the format: (index, x, y, w, h)
boxes.sort(key=lambda box: box[2])  # Sort by the 'y' value (third column)

# Prompt the user for the index of the box to highlight
user_input = input("Enter the box index to highlight: ")
try:
    box_index = int(user_input)
except ValueError:
    print("Error: Invalid input. Please enter a valid box index.")
    exit()

# Find the box data for the given index
box_data = next((box for box in boxes if box[0] == box_index), None)
if not box_data:
    print(f"Error: Box with index {box_index} not found in the Excel file.")
    exit()

# Extract box dimensions
_, x, y, w, h = box_data

# Highlight the selected box with a unique color (e.g., bright pink)
highlight_color = (255, 0, 255)  # Bright pink
cv2.rectangle(image, (int(x), int(y)), (int(x + w), int(y + h)), highlight_color, 20)  # Thicker border for visibility

# Display the image with the highlighted box
cv2.imshow(f"Highlighted Box {box_index}", image)
cv2.waitKey(0)
cv2.destroyAllWindows()

# Save the image with the highlighted box
output_path = os.path.join(os.path.dirname(__file__), '../book/static/book/images/cbf-layout1_highlighted_box.jpg')  # Update with your desired output path
cv2.imwrite(output_path, image)

print(f"Box {box_index} highlighted and saved to {output_path}")
